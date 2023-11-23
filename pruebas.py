import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
import time

# Abre el archivo Excel y selecciona la hoja
workbook = openpyxl.load_workbook('registros.xlsx')
sheet = workbook['Sheet']

# Inicializa el controlador del navegador (asegúrate de tener el controlador correspondiente instalado)
chrome_service = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
driver.maximize_window()
driver.get(url='https://localhost:7244/usuarios/registrar') 

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    curp = row[0].value
    nombre = row[1].value
    primer_apellido = row[2].value
    segundo_apellido = row[3].value
    fecha_nacimiento = row[4].value
    genero = row[5].value
    pais = row[6].value  
    tipo_documento = row[7].value
    codigo_postal = row[8].value
    colonia = row[9].value
    calle = row[10].value
    numero_exterior = row[11].value
    entre_calles = row[12].value
    correo_electronico = row[13].value
    confirmar_correo = row[14].value
    contrasena = row[15].value
    lada = row[16].value
    numero_celular = row[17].value
  

    if curp is None:
        time.sleep(7)
        checkbox = driver.find_element(By.CSS_SELECTOR, "#form > div:nth-child(2) > div.w-100.w-sm-75.mx-auto > div.steps > div.step-1 > div > div:nth-child(4) > div > div > div.col-sm-4.order-0.order-sm-2.text-start.text-sm-center > div > div.col-sm-12.col-5.text-end.text-sm-center > span > span > span.k-switch-thumb-wrap > span")
        checkbox.click()
        # Completa los campos del formulario
        driver.find_element(By.CSS_SELECTOR, "#nombre").send_keys(nombre)
        driver.find_element(By.CSS_SELECTOR, "#primerApellido").send_keys(primer_apellido)
        driver.find_element(By.CSS_SELECTOR, "#segundoApellido").send_keys(segundo_apellido)
        driver.find_element(By.CSS_SELECTOR, "#fechaNacimiento").send_keys(fecha_nacimiento)

        # Selecciona el género en un menú desplegable
        genero_dropdown = Select(driver.find_element(By.CSS_SELECTOR, "#genero"))
        genero_dropdown.select_by_index(genero)

        pais_dropdown = Select(driver.find_element(By.CSS_SELECTOR, "#paisId"))
        pais_dropdown.select_by_index(pais)

        tipo_documento_dropdown = Select(driver.find_element(By.CSS_SELECTOR, "#tipoDocumento"))
        tipo_documento_dropdown.select_by_index(tipo_documento)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        cp = driver.find_element(By.CSS_SELECTOR, "#codigoPostal").send_keys(codigo_postal)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        div = driver.find_element(By.CSS_SELECTOR,"#infoDomicilio > div > div")
        time.sleep(3)
        div.click()
        time.sleep(10)
        # Encuentra la lista desplegable
        colonia_select = driver.find_element(By.CSS_SELECTOR, "#SelectColoniaId > div > span > span.pc-element.pc-trigger").click()
        opciones = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".picker li")))
        for option in opciones:
            if option.text == colonia:
                time.sleep(5)
                option.click()
                break
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "#calle").send_keys(calle)
        driver.find_element(By.CSS_SELECTOR, "#numeroExterior").send_keys(numero_exterior)
        driver.find_element(By.CSS_SELECTOR, "#entreCalles").send_keys(entre_calles)
        driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
        time.sleep(2)

        # Completa los campos de correo electrónico y número de celular
        driver.find_element(By.CSS_SELECTOR, "#correo").send_keys(correo_electronico)
        driver.find_element(By.CSS_SELECTOR, "#confirmarCorreo").send_keys(confirmar_correo)

        # Encuentra el elemento select por su nombre
        lada_select = Select(driver.find_element(By.CSS_SELECTOR, "#lada"))
        lada_select.select_by_index(lada)

        driver.find_element(By.CSS_SELECTOR, "#celular").send_keys(numero_celular)
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
        time.sleep(2)

        # Agrega los campos de Contraseña y Confirmar Contraseña
        driver.find_element(By.CSS_SELECTOR, "#contrasena").send_keys(contrasena)
        driver.find_element(By.CSS_SELECTOR, "#confirmarContrasena").send_keys(contrasena)
        time.sleep(5)

        # Marca el checkbox "Acepto Términos y Condiciones"
        driver.find_element(By.CSS_SELECTOR, '#form > div:nth-child(2) > div.w-100.w-sm-75.mx-auto > div.steps > div.step-3 > div > div.col-sm-12 > div > p > span > span.k-switch-thumb-wrap > span').click()
        time.sleep(2)

        # Envía el formulario
        driver.find_element(By.CSS_SELECTOR, "#finalizarBtn").click()
        time.sleep(15)
        driver.find_element(By.CSS_SELECTOR, "body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-actions > button.swal2-confirm.swal2-styled").click()
        time.sleep(7)
        driver.find_element(By.CSS_SELECTOR, "#content > div > div > div > div:nth-child(1) > div > div > div.row > div:nth-child(2) > a").click()

        # Espera a que la página se cargue antes de continuar al siguiente registro
        driver.implicitly_wait(10)
    else:
        driver.find_element(By.CSS_SELECTOR, '#curp').send_keys(curp)
        # Esperar a que el marco esté disponible y cambiar a él
        frame_locator = (By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")
        WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it(frame_locator))
        # Esperar a que el cuadro de la casilla de verificación sea clickable y hacer clic en él
        checkbox_locator = (By.CSS_SELECTOR, "div.recaptcha-checkbox-border")
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(checkbox_locator)).click()
        driver.switch_to.default_content()
        # Ahora intenta hacer clic en el botón
        driver.find_element(By.CSS_SELECTOR, "#buscarBtn").click()
        time.sleep(10)
        try:
            if driver.find_element(By.CSS_SELECTOR,'body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-header'):
                driver.find_element(By.CSS_SELECTOR, 'body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-actions > button.swal2-confirm.swal2-styled').click()
                # Completa los campos del formulario
                driver.find_element(By.CSS_SELECTOR, "#nombre").send_keys(nombre)
                driver.find_element(By.CSS_SELECTOR, "#primerApellido").send_keys(primer_apellido)
                driver.find_element(By.CSS_SELECTOR, "#segundoApellido").send_keys(segundo_apellido)
                driver.find_element(By.CSS_SELECTOR, "#fechaNacimiento").send_keys(fecha_nacimiento)

                # Selecciona el género en un menú desplegable
                genero_dropdown = Select(driver.find_element(By.CSS_SELECTOR, "#genero"))
                genero_dropdown.select_by_index(genero)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                cp = driver.find_element(By.CSS_SELECTOR, "#codigoPostal").send_keys(codigo_postal)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                div = driver.find_element(By.CSS_SELECTOR,"#infoDomicilio > div > div")
                time.sleep(3)
                div.click()
                time.sleep(10)
                # Encuentra la lista desplegable
                colonia_select = driver.find_element(By.CSS_SELECTOR, "#SelectColoniaId > div > span > span.pc-element.pc-trigger").click()
                opciones = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".picker li")))
                for option in opciones:
                    if option.text == colonia:
                        time.sleep(5)
                        option.click()
                        break
                time.sleep(3)
                driver.find_element(By.CSS_SELECTOR, "#calle").send_keys(calle)
                driver.find_element(By.CSS_SELECTOR, "#numeroExterior").send_keys(numero_exterior)
                driver.find_element(By.CSS_SELECTOR, "#entreCalles").send_keys(entre_calles)
                driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
                time.sleep(2)

                # Completa los campos de correo electrónico y número de celular
                driver.find_element(By.CSS_SELECTOR, "#correo").send_keys(correo_electronico)
                driver.find_element(By.CSS_SELECTOR, "#confirmarCorreo").send_keys(confirmar_correo)

                # Encuentra el elemento select por su nombre
                lada_select = Select(driver.find_element(By.CSS_SELECTOR, "#lada"))
                lada_select.select_by_index(lada)

                driver.find_element(By.CSS_SELECTOR, "#celular").send_keys(numero_celular)
                time.sleep(3)
                driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
                time.sleep(2)

                # Agrega los campos de Contraseña y Confirmar Contraseña
                driver.find_element(By.CSS_SELECTOR, "#contrasena").send_keys(contrasena)
                driver.find_element(By.CSS_SELECTOR, "#confirmarContrasena").send_keys(contrasena)
                time.sleep(5)

                # Marca el checkbox "Acepto Términos y Condiciones"
                driver.find_element(By.CSS_SELECTOR, '#form > div:nth-child(2) > div.w-100.w-sm-75.mx-auto > div.steps > div.step-3 > div > div.col-sm-12 > div > p > span > span.k-switch-thumb-wrap > span').click()
                time.sleep(2)

                # Envía el formulario
                driver.find_element(By.CSS_SELECTOR, "#finalizarBtn").click()
                time.sleep(15)
                driver.find_element(By.CSS_SELECTOR, "body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-actions > button.swal2-confirm.swal2-styled").click()
                time.sleep(7)
                driver.find_element(By.CSS_SELECTOR, "#content > div > div > div > div:nth-child(1) > div > div > div.row > div:nth-child(2) > a").click()

                # Espera a que la página se cargue antes de continuar al siguiente registro
                driver.implicitly_wait(10)

        except NoSuchElementException:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            cp = driver.find_element(By.CSS_SELECTOR, "#codigoPostal").send_keys(codigo_postal)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            div = driver.find_element(By.CSS_SELECTOR,"#infoDomicilio > div > div")
            time.sleep(3)
            div.click()
            time.sleep(10)
            # Encuentra la lista desplegable
            colonia_select = driver.find_element(By.CSS_SELECTOR, "#SelectColoniaId > div > span > span.pc-element.pc-trigger").click()
            opciones = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".picker li")))
            for option in opciones:
                if option.text == colonia:
                    time.sleep(5)
                    option.click()
                    break
            time.sleep(3)
            driver.find_element(By.CSS_SELECTOR, "#calle").send_keys(calle)
            driver.find_element(By.CSS_SELECTOR, "#numeroExterior").send_keys(numero_exterior)
            driver.find_element(By.CSS_SELECTOR, "#entreCalles").send_keys(entre_calles)
            driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
            time.sleep(2)

            # Completa los campos de correo electrónico y número de celular
            driver.find_element(By.CSS_SELECTOR, "#correo").send_keys(correo_electronico)
            driver.find_element(By.CSS_SELECTOR, "#confirmarCorreo").send_keys(confirmar_correo)

            # Encuentra el elemento select por su nombre
            lada_select = Select(driver.find_element(By.CSS_SELECTOR, "#lada"))
            lada_select.select_by_index(lada)

            driver.find_element(By.CSS_SELECTOR, "#celular").send_keys(numero_celular)
            time.sleep(3)
            driver.find_element(By.CSS_SELECTOR, "#siguienteBtn").click()
            time.sleep(2)

            # Agrega los campos de Contraseña y Confirmar Contraseña
            driver.find_element(By.CSS_SELECTOR, "#contrasena").send_keys(contrasena)
            driver.find_element(By.CSS_SELECTOR, "#confirmarContrasena").send_keys(contrasena)
            time.sleep(5)

            # Marca el checkbox "Acepto Términos y Condiciones"
            driver.find_element(By.CSS_SELECTOR, '#form > div:nth-child(2) > div.w-100.w-sm-75.mx-auto > div.steps > div.step-3 > div > div.col-sm-12 > div > p > span > span.k-switch-thumb-wrap > span').click()
            time.sleep(2)

            # Envía el formulario
            driver.find_element(By.CSS_SELECTOR, "#finalizarBtn").click()
            time.sleep(15)
            driver.find_element(By.CSS_SELECTOR, "body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-actions > button.swal2-confirm.swal2-styled").click()
            time.sleep(7)
            driver.find_element(By.CSS_SELECTOR, "#content > div > div > div > div:nth-child(1) > div > div > div.row > div:nth-child(2) > a").click()

            # Espera a que la página se cargue antes de continuar al siguiente registro
            driver.implicitly_wait(10)

if driver:
    # Cierra el navegador cuando hayas terminado
    driver.quit()
